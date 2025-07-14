from flask import Flask, request, send_file, jsonify
import os
import tempfile
import pdfplumber
import pandas as pd
import pytesseract
from pdf2image import convert_from_path
from PIL import Image
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import openpyxl
import re

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def extract_from_pdfplumber(pdf_path):
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                for row in table:
                    if row and any(cell is not None and str(cell).strip() for cell in row):
                        rows.append(row)
    return rows

def extract_from_ocr(pdf_path):
    text_data = ""
    images = convert_from_path(pdf_path)
    for image in images:
        text_data += pytesseract.image_to_string(image)
    return extract_rows_from_text(text_data)

def extract_rows_from_text(text):
    rows = []
    lines = text.splitlines()
    for line in lines:
        match = re.match(r"(\d{2}-\d{2}-\d{4})\s+(.+?)\s+([\d,]+\.\d{2})\((Dr|Cr)\)\s+([\d,]+\.\d{2})\((Dr|Cr)\)", line)
        if match:
            date, narration, amount, amt_type, balance, bal_type = match.groups()
            amount_with_type = f"{amount} ({amt_type})"
            rows.append([date, narration.strip(), amount_with_type, balance])
    return rows

def save_to_excel(rows, headers=None):
    wb = Workbook()
    ws = wb.active
    if headers:
        ws.append(headers)
    for row in rows:
        ws.append(row)
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 5
    output_path = os.path.join(tempfile.gettempdir(), 'converted.xlsx')
    wb.save(output_path)
    return output_path

def auto_resize_excel_columns(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for column_cells in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 5
    wb.save(path)

@app.route('/upload', methods=['POST'])
def upload_pdf():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '' or not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Please upload a valid PDF file'}), 400

    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    try:
        # Try structured text first
        rows = extract_from_pdfplumber(filepath)
        if rows and len(rows) > 1:
            df = pd.DataFrame(rows[1:], columns=rows[0])
            excel_path = os.path.join(tempfile.gettempdir(), 'converted.xlsx')
            df.to_excel(excel_path, index=False)
            auto_resize_excel_columns(excel_path)
            return send_file(excel_path, as_attachment=True, download_name="converted.xlsx")

        # Try OCR fallback
        rows = extract_from_ocr(filepath)
        if rows:
            excel_path = save_to_excel(rows, headers=["Date", "Narration", "Amount", "Balance"])
            return send_file(excel_path, as_attachment=True, download_name="converted.xlsx")

        return jsonify({'error': 'No tabular data found'}), 400

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/', methods=['GET'])
def index():
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>PDF to Excel Converter</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                padding: 40px;
                background-color: #f7f7f7;
            }
            form {
                background: #fff;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 0 10px rgba(0,0,0,0.1);
                max-width: 400px;
                margin: auto;
            }
            h2 {
                text-align: center;
            }
            input[type="file"] {
                display: block;
                margin: 20px 0;
                width: 100%;
            }
            input[type="submit"] {
                background: #4CAF50;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 4px;
                cursor: pointer;
            }
            input[type="submit"]:hover {
                background: #45a049;
            }
        </style>
    </head>
    <body>
        <form action="/upload" method="post" enctype="multipart/form-data">
            <h2>Upload PDF Bank Statement</h2>
            <input type="file" name="file" accept=".pdf" required>
            <input type="submit" value="Convert to Excel">
        </form>
    </body>
    </html>
    '''

if __name__ == '__main__':
    app.run(debug=True, port=5000)


